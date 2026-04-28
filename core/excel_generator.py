"""Genera el Excel adjunto con las glosas a corregir para un trabajo.

La columna `glosa_enviada` usa CellRichText para marcar en rojo el
caracter o segmento problemático. La columna `error_detectado` describe
la regla rota + posición. Sin asteriscos ni underscores ni otros
decoradores textuales: solo color, o nada (cuando el error no admite
marcado puntual).

Las glosas rotas y sus segmentos a marcar provienen de
`core.simulador.generar_glosa_con_error`, que cubre errores en todos
los componentes (IOC, RUT, ID, iniciales, orden).
"""

import io
import random
from datetime import datetime, timedelta
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Border, Font, PatternFill, Side

from core.simulador import generar_glosa_con_error

HEADERS = (
    "glosa_enviada",
    "monto",
    "fecha",
    "error_detectado",
    "glosa_correcta",
)

_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="FF404040")
_GRAY_FILL = PatternFill("solid", fgColor="FFEFEFEF")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFFFF")
_BORDER = Border(
    left=Side(style="thin", color="FF888888"),
    right=Side(style="thin", color="FF888888"),
    top=Side(style="thin", color="FF888888"),
    bottom=Side(style="thin", color="FF888888"),
)

_ROJO = InlineFont(color="FFFF0000")


def _construir_rich_text(
    glosa: str, segmentos_rojos: list[tuple[int, int]]
) -> CellRichText:
    """Convierte (glosa, ranges) a CellRichText con TextBlock rojos."""
    if not segmentos_rojos:
        return CellRichText([glosa])

    partes: list[Any] = []
    cursor = 0
    for start, end in sorted(segmentos_rojos):
        if start > cursor:
            partes.append(glosa[cursor:start])
        partes.append(TextBlock(_ROJO, glosa[start:end]))
        cursor = end
    if cursor < len(glosa):
        partes.append(glosa[cursor:])
    return CellRichText(partes)


def _aplicar_estilo(cell, fill: PatternFill) -> None:
    cell.fill = fill
    cell.border = _BORDER


def generar_excel_adjunto(trabajo: dict[str, Any]) -> tuple[bytes, str]:
    """Devuelve (excel_bytes, nombre_archivo) con 5 glosas rotas a corregir."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Glosas a corregir"

    # Header
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        _aplicar_estilo(cell, _HEADER_FILL)

    today = datetime.now().date()
    for row_idx in range(2, 7):
        err = generar_glosa_con_error()
        rich = _construir_rich_text(err["glosa"], err["rojos"])
        monto = random.randint(100_000, 5_000_000)
        fecha = (today - timedelta(days=random.randint(0, 30))).strftime("%Y-%m-%d")

        c = ws.cell(row=row_idx, column=1, value=rich)
        _aplicar_estilo(c, _GRAY_FILL)
        c = ws.cell(row=row_idx, column=2, value=monto)
        _aplicar_estilo(c, _GRAY_FILL)
        c = ws.cell(row=row_idx, column=3, value=fecha)
        _aplicar_estilo(c, _GRAY_FILL)
        c = ws.cell(row=row_idx, column=4, value=err["error_desc"])
        _aplicar_estilo(c, _GRAY_FILL)
        c = ws.cell(row=row_idx, column=5, value=None)
        _aplicar_estilo(c, _WHITE_FILL)

    for col, w in {"A": 50, "B": 12, "C": 12, "D": 44, "E": 42}.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)

    safe_desc = (
        trabajo["descripcion"][:30]
        .replace(" ", "_")
        .replace("/", "_")
        .replace(":", "")
    )
    filename = f"glosas_a_corregir_t{trabajo['trabajo_id']}_{safe_desc}.xlsx"
    return buf.getvalue(), filename
